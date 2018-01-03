import MySQLdb
import smtplib
import time
from time import strftime
from openpyxl import load_workbook
import sys
reload(sys)
sys.setdefaultencoding('utf-8')
wb=load_workbook('UberClients.xlsx')
ws = wb.active

#Database Connection
conn = MySQLdb.connect(host="localhost", user="root", passwd="", db="valdatta")
cursor = conn.cursor()


for x in range(1,10):
    name = ws.cell(row=1,column=x).value
    id_unico = str(ws.cell(row=2,column=x).value)
    curp = str(ws.cell(row=3,column=x).value)
    rfc = str(ws.cell(row=4,column=x).value)
    birthdate = str(ws.cell(row=5,column=x).value)
    #birthdate = birthdate.split('/',2)
    #print(birthdate)
    #birthdate = birthdate[2]+ " " + birthdate[0] + " " + birthdate[1]
    #birthdate = time.strftime('%Y-%m-%d %H:%M',time.strptime(birthdate,"%Y %m %d"))
    address = str(ws.cell(row=6,column=x).value)
    cellphone = str(ws.cell(row=7,column=x).value)
    educationLevel = str(ws.cell(row=8,column=x).value)
    dependendants = str(ws.cell(row=9,column=x).value)
    livingStatus = str(ws.cell(row=10,column=x).value)
    credit = str(ws.cell(row=11,column=x).value)
    creditCard = str(ws.cell(row=12,column=x).value)
    lastDigits = str(ws.cell(row=13,column=x).value)
    business = ws.cell(row=14,column=x).value
    businessField = ws.cell(row=15,column=x).value
    businessRFC = str(ws.cell(row=16,column=x).value)
    businessAddress = ws.cell(row=17,column=x).value
    businessFoundation = str(ws.cell(row=18,column=x).value)
    #businessFoundation = businessFoundation.split('/',2)
    #businessFoundation = businessFoundation[2]+ " " + businessFoundation[0] + " " + businessFoundation[1]
    #businessFoundation = time.strftime('%Y-%m-%d %H:%M',time.strptime(businessFoundation,"%Y %m %d"))
    brand = ws.cell(row=19,column=x).value
    businessPhone = str(ws.cell(row=20,column=x).value)
    businessCellphone = str(ws.cell(row=21,column=x).value)
    webSite = str(ws.cell(row=22,column=x).value)
    partnersNumber = str(ws.cell(row=23,column=x).value)
    employeesNumber = str(ws.cell(row=24,column=x).value)
    isPartner = str(ws.cell(row=25,column=x).value)
    sellsType = str(ws.cell(row=26,column=x).value)
    operationDays = str(ws.cell(row=27,column=x).value)
    ref1Name = str(ws.cell(row=28,column=x).value)
    ref1Contact = str(ws.cell(row=29,column=x).value)
    ref1Phone = str(ws.cell(row=30,column=x).value)
    ref2Name = str(ws.cell(row=31,column=x).value)
    ref2Contact = str(ws.cell(row=32,column=x).value)
    ref2Phone = str(ws.cell(row=33,column=x).value)
    ref3Name = str(ws.cell(row=34,column=x).value)
    ref3Contact = str(ws.cell(row=35,column=x).value)
    ref3Phone = str(ws.cell(row=36,column=x).value)
    bank = str(ws.cell(row=37,column=x).value)
    cardNumber = str(ws.cell(row=38,column=x).value)
    clabe = str(ws.cell(row=39,column=x).value)
    CapitalToPay = str(ws.cell(row=40,column=x).value)
    InterestsToPay = str(ws.cell(row=41,column=x).value)
    interestRate = str(ws.cell(row=42,column=x).value)
    amount = float(CapitalToPay) + float(InterestsToPay)
    arrearRate = interestRate
    periodicity = 1
    term =1
    creditStart = str(ws.cell(row=43,column=x).value)
    ArrearToPay = 0

    print('Nombre:' + str(name))


    cursor.execute('INSERT INTO clients (name,curp,rfc,birthDay,address,phone,cellphone,bank,clabe,created_at,updated_at,marritalStatus,educationLevel,dependendants,livingStatus,credit,creditCard,lastDigits,business,businessField,businessRFC,businessAddress,businessFundation,brand,businessPhone,businessCellphone,webSite,partnersNumber,employeesNumbers,isPartner,sellsType,operationDays,ref1Name,ref1Contact,ref1Phone,ref2Name,ref2Contact,ref2Phone,ref3Name,ref3Contact,ref3Phone,cardNumber) values (\"'+name+'\",\"'+curp+'\",\"'+rfc+'\",\"'+birthdate+'\",\"'+address+'\",\"'+cellphone+'\",\"'+cellphone+'\",\"'+bank+'\",\"'+clabe+'\","2017-12-12","2017-12-12","1",\"'+educationLevel+'\",\"'+dependendants+'\",\"'+livingStatus+'\",\"'+credit+'\",\"'+creditCard+'\",\"'+lastDigits+'\",\"'+business+'\",\"'+businessField+'\",\"'+businessRFC+'\",\"'+businessAddress+'\",\"'+businessFoundation+'\","1","1","1","1","1","1","1","1","1","1","1","1","1","1","3","1","1","3","151515151561665")')
    conn.commit()

    cursor.execute('SELECT LAST_INSERT_ID()')
    client_id = cursor.fetchone()

    #Create Credit for CreditAccounts
    cursor.execute('INSERT INTO credits (client_id,type,status,amount,term,periodicity,interestRate,arrearRate,realRate,creditStart,CapitalToPay,InterestsToPay,ArrearToPay,applications_id,created_at,updated_at,FinancialProduct_id) values (\"'+str(client_id[0])+'\","3","1",\"'+str(amount)+'\",\"'+str(term)+'\",\"'+str(periodicity)+'\",\"'+str(interestRate)+'\",\"'+str(arrearRate)+'\","0",\"'+str(creditStart)+'\",\"'+str(CapitalToPay)+'\",\"'+str(InterestsToPay)+'\",\"'+str(ArrearToPay)+'\","0",\"'+str(creditStart)+'\",\"'+str(creditStart)+'\","0")' )
    conn.commit()

    cursor.execute('SELECT LAST_INSERT_ID()')
    credit_id = cursor.fetchone()

    #Create Credit Accounts and Payment Accounts

    cursor.execute('INSERT INTO paymentaccounts (created_at,updated_at,client_id,status,type,id_unico) values ("2017-12-12","2017-12-12",\"'+str(client_id[0])+'\","1","1",\"'+str(id_unico)+'\")')
    conn.commit()

    cursor.execute('SELECT LAST_INSERT_ID()')
    paymentaccounts = cursor.fetchone()

    cursor.execute('INSERT INTO creditaccounts (created_at,updated_at,credit_id,PaymentAccounts_id) values ("2017-12-12","2017-12-12",\"'+str(credit_id[0])+'\",\"'+str(paymentaccounts[0])+'\")')

#row = cursor.fetchone()

conn.close()

#print(row)
