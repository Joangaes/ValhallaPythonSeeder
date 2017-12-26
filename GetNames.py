from openpyxl import load_workbook
import sys
import csv
from xlsxwriter.workbook import Workbook
reload(sys)
sys.setdefaultencoding('utf-8')
wb=load_workbook('171212 Gastos Resumenes.xlsx')
ws = wb.active

names=[]
nombre = ws.cell(row=2,column=5).value
names.append(nombre);

for x in range(3,2184):
    nombre = ws.cell(row=x,column=5).value
    #print(len(names))
    for y in range(0,len(names)):
        #print(nombre)
        if(nombre==names[y]):
            break
        if(y==len(names)-1):
            names.append(nombre)
workbook = Workbook('NombresClientes.xlsx')
worksheet = workbook.add_worksheet()
for y in range(0,len(names)):
    worksheet.write(y+1, 1, names[y])
    print(names[y])
workbook.close()
