import MySQLdb
import smtplib
from openpyxl import load_workbook
import sys
reload(sys)
sys.setdefaultencoding('utf-8')

def Colocacion(fecha_originacion,Coloca,client_id,conn,cursor):
    cursor.execute("""SELECT * FROM credits WHERE client_id = %s AND type=%s""", (client_id[0],"2",))
    credit_id = cursor.fetchone()
    print('Col')
    cursor.execute('INSERT INTO expectedcreditspayments (created_at,updated_at,Start_date,amount,credit_id) values (\"'+str(fecha_originacion)+'\",\"'+str(fecha_originacion)+'\",\"'+str(fecha_originacion)+'\",\"'+str(Coloca)+'\",\"'+str(credit_id[0])+'\")')
    conn.commit()

def Pago(fecha_originacion,fecha_esperada_pago,fecha_pago_real,Recupera,Intereses,IVA,Mora,IVAMora,PagoTot,nombrecliente):
    print('Pago')
    cursor.execute("""SELECT * FROM credits WHERE client_id = %s AND type=%s""", (client_id[0],"2",))
    credit_id = cursor.fetchone()
    cursor.execute("""SELECT * FROM ledgers WHERE company = %s AND amount=%s""", (nombrecliente,PagoTot,))
    ledger_id = cursor.fetchone()
    print('fecha esperada de pago: ' + str(fecha_esperada_pago))
    if ledger_id is None:
        return
    else:
        cursor.execute('INSERT INTO payments (created_at,updated_at,credit_id,ledger_id,amount,interest,interestIVA,arrear,arrearIVA,expectedDate,realDate,type) values (\"'+str(fecha_pago_real)+'\",\"'+str(fecha_pago_real)+'\",\"'+str(credit_id[0])+'\",\"'+str(ledger_id[0])+'\",\"'+str(PagoTot)+'\",\"'+str(Intereses)+'\",\"'+str(IVA)+'\",\"'+str(Mora)+'\",\"'+str(IVAMora)+'\",\"'+str(fecha_esperada_pago)+'\",\"'+str(fecha_pago_real)+'\","1")')
        conn.commit()



wb=load_workbook('171207 Cartera Factoraje.xlsx', data_only=True)
ws = wb['BD Factoraje']
conn = MySQLdb.connect(host="localhost", user="root", passwd="", db="valdatta")
cursor = conn.cursor()





for x in range(2,404):
    print(x)
    nombrecliente = ws.cell(row=x,column=2).value
    cursor.execute("""SELECT * FROM clients WHERE name = %s""", (nombrecliente,))
    client_id = cursor.fetchone()
    fecha_originacion = ws.cell(row=x,column=5).value
    print(fecha_originacion)
    Pago_o_Paga = str(ws.cell(row=x,column=8).value)
    fecha_esperada_pago = ws.cell(row=x,column=9).value
    print('Pago o Paga: '+ str(Pago_o_Paga))
    if(Pago_o_Paga=='1'):
        fecha_pago_real = ws.cell(row=x,column=12).value
        Recupera = ws.cell(row=x,column=18).value
        Intereses = ws.cell(row=x,column=19).value
        IVA = ws.cell(row=x,column=20).value
        Mora = ws.cell(row=x,column=21).value
        if Mora is None:
            Mora = 0
            IVAMora = 0
        else:
            IVAMora = ws.cell(row=x,column=22).value
            PagoTot = ws.cell(row=x,column=23).value
            Pago(fecha_originacion,fecha_esperada_pago,fecha_pago_real,Recupera,Intereses,IVA,Mora,IVAMora,PagoTot,nombrecliente)
    else:
        if(Pago_o_Paga=='0'):
            Coloca = ws.cell(row=x,column=17).value
            Colocacion(fecha_originacion,Coloca,client_id,conn,cursor)
