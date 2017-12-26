import MySQLdb
import smtplib
from openpyxl import load_workbook
import sys
reload(sys)
sys.setdefaultencoding('utf-8')
wb=load_workbook('171212 Gastos Resumenes.xlsx')
ws = wb.active

#Database Connection
conn = MySQLdb.connect(host="localhost", user="root", passwd="", db="valdatta")
cursor = conn.cursor()


for x in range(2,2184):
    generalConcept = ws.cell(row=x,column=4).value
    concept = str(ws.cell(row=x,column=5).value)
    bank=str(ws.cell(row=x,column=6).value)
    amount = str(ws.cell(row=x,column=7).value)
    comments = str(ws.cell(row=x,column=2).value)
    date = str(ws.cell(row=x,column=3).value)

    print(concept)

    cursor.execute('INSERT INTO ledgers (created_at,updated_at,date,amount,iva,concept,total,bank,company,generalConcept,movementType,document,comments) values ("2017-12-12","2017-12-12",\"'+date+'\",\"'+amount+'\","0",\"'+concept+'\",\"'+amount+'\",\"'+bank+'\",\"'+concept+'\",\"'+generalConcept+'\","1","",\"'+comments+'\")')
    conn.commit()
#row = cursor.fetchone()

conn.close()

#print(row)
