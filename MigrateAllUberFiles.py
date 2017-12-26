import MySQLdb
import smtplib
import time
from time import strftime
from openpyxl import load_workbook
import sys
reload(sys)
sys.setdefaultencoding('utf-8')
import glob

conn = MySQLdb.connect(host="localhost", user="root", passwd="", db="valdatta")
cursor = conn.cursor()



files = []
files = glob.glob("Uber_To_Arkafin/*.xlsx")

for x in range (0,len(files)):

    wb=load_workbook(files[x])
    ws = wb.active
    for y in range (2,6):
        uber_user_uuid = str(ws.cell(row=y,column=2).value)
        amount_charged = str(ws.cell(row=y,column=4).value)
        due_date = str(ws.cell(row=y,column=6).value)
        due_date = due_date.split('/',2)
        due_date = due_date[2]+ " " + due_date[0] + " " + due_date[1]
        due_date = time.strftime('%Y-%m-%d %H:%M',time.strptime(due_date,"%Y %m %d"))
        #print(due_date)
        generalConcept = str(ws.cell(row=y,column=7).value)
        concept = str(ws.cell(row=y,column=7).value)
        bank = '3'
        company = 'Visor'
        #query = 'SELECT * FROM PaymentAccounts WHERE id_unico = ' + str(uber_user_uuid)
        cursor.execute("""SELECT * FROM PaymentAccounts WHERE id_unico = %s""", (uber_user_uuid,))
        PaymentAccounts = cursor.fetchone()
        print('ID PaymentAccounts: ' + str(PaymentAccounts[0]))
        cursor.execute("""SELECT * FROM CreditAccounts WHERE PaymentAccounts_id = %s""", (PaymentAccounts[0],))
        credit_id = cursor.fetchone()
        cursor.execute('INSERT INTO ledgers (created_at,updated_at,date,amount,iva,concept,total,bank,company,generalConcept,movementType,document,comments,type) values ("2017-12-12","2017-12-12",\"'+due_date+'\",\"'+amount_charged+'\","0",\"'+concept+'\",\"'+amount_charged+'\",\"'+bank+'\",\"'+company+'\",\"'+generalConcept+'\","1","","","1")')
        conn.commit()
        cursor.execute('SELECT LAST_INSERT_ID()')
        lastidLedgers = cursor.fetchone()
        cursor.execute("""SELECT * FROM Credits WHERE id = %s""", (credit_id[3],))
        credit = cursor.fetchone()
        interest = float(amount_charged) * float(credit[7])/12 * float(credit[9])
        interestIVA = interest*.16
        amountleft =float(amount_charged) - interest - interestIVA
        cursor.execute('INSERT INTO Payments (created_at,updated_at,credit_id,ledger_id,amount,interest,interestIVA,arrear,arrearIVA,expectedDate,realDate,type) values ("2017-12-12","2017-12-12",\"'+str(credit[0])+'\",\"'+str(lastidLedgers[0])+'\",\"'+str(amountleft)+'\",\"'+str(interest)+'\",\"'+str(interestIVA)+'\","0","0",\"'+due_date+'\",\"'+due_date+'\","1")')



        model = cursor.fetchone()
        print(credit)
        print('ID Credito: '+ str(credit_id[0]))


conn.close()
