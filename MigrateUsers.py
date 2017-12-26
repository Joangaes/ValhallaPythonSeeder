import MySQLdb
import smtplib
from openpyxl import load_workbook
import sys
reload(sys)
sys.setdefaultencoding('utf-8')
wb=load_workbook('171207 TODO EN UNO.xlsx', data_only=True)
ws = wb['BASE DE DATOS LEGALES']
#print(ws.cell(row=,column=1).value)

conn = MySQLdb.connect(host="localhost", user="root", passwd="", db="valdatta")
cursor = conn.cursor()

for x in range(2,101):
    name = ws.cell(row=x,column=5).value
    address = str(ws.cell(row=x,column=6).value)
    rfc=str(ws.cell(row=x,column=13).value)
    bank = str(ws.cell(row=x,column=14).value)
    clabe = str(ws.cell(row=x,column=15).value)
    phone = str(ws.cell(row=x,column=11).value)

    print(bank)

    cursor.execute('INSERT INTO clients (name,curp,rfc,birthDay,address,phone,cellphone,bank,clabe,created_at,updated_at,marritalStatus,educationLevel,dependendants,livingStatus,credit,creditCard,lastDigits,business,businessField,businessRFC,businessAddress,businessFundation,brand,businessPhone,businessCellphone,webSite,partnersNumber,employeesNumbers,isPartner,sellsType,operationDays,ref1Name,ref1Contact,ref1Phone,ref2Name,ref2Contact,ref2Phone,ref3Name,ref3Contact,ref3Phone,cardNumber) values (\"'+name+'\","v54645",\"'+rfc+'\","1996-12-01",\"'+address+'\","0",\"'+phone+'\",\"'+bank+'\",\"'+clabe+'\","2017-12-12","2017-12-12","1","1","1","1","1","1","1","1","1","1","1","2017-12-12","1","1","1","1","1","1","1","1","1","1","1","1","1","1","3","1","1","3","151515151561665")')
    conn.commit()
#row = cursor.fetchone()

conn.close()

#print(row)
