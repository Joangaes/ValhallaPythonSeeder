import MySQLdb
import smtplib
from openpyxl import load_workbook
import sys
reload(sys)
sys.setdefaultencoding('utf-8')
wb=load_workbook('171207 TODO EN UNO.xlsx', data_only=True)
#print(wb.sheetnames)
ws = wb['PRESTAMOS']
conn = MySQLdb.connect(host="localhost", user="root", passwd="", db="valdatta")
cursor = conn.cursor()
print(ws.cell(row=3,column=12).value)
for x in range(2,354):
    nombrecliente = ws.cell(row=x,column=5).value
    print(nombrecliente + '    -   ' + str(x))
    amount = str(ws.cell(row=x,column=10).value)
    tipo = ws.cell(row=x,column=6).value
    if(tipo == 'CS'):
        tipo='1';
    else:
        tipo='2'
    term = str(ws.cell(row=x,column=11).value)
    if(term=='None'):
        term='12'
    interestRate=str(ws.cell(row=x,column=12).value)
    if(interestRate=='None'):
        interestRate=0
    else:
        #print(interestRate)
        interestRate=str(float(interestRate)/100)
        #print('interes2')
    periodicity = '1'
    arrearRate = interestRate
    creditStart= ws.cell(row=x,column=1).value
    CapitalToPay = amount
    InterestsToPay = str(float(amount) * float(interestRate))
    ArrearToPay='0'
    applications_id = '0';
    #print(term)
    cursor.execute("""SELECT * FROM clients WHERE name = %s""", (nombrecliente,))
    client_id = cursor.fetchone()
    print(client_id)
    if(client_id!=None):
        cursor.execute('INSERT INTO credits (client_id,type,status,amount,term,periodicity,interestRate,arrearRate,realRate,creditStart,CapitalToPay,InterestsToPay,ArrearToPay,applications_id,created_at,updated_at,FinancialProduct_id) values (\"'+str(client_id[0])+'\",\"'+str(tipo)+'\","1",\"'+str(amount)+'\",\"'+str(term)+'\",\"'+str(periodicity)+'\",\"'+str(interestRate)+'\",\"'+str(arrearRate)+'\","0",\"'+str(creditStart)+'\",\"'+str(CapitalToPay)+'\",\"'+str(InterestsToPay)+'\",\"'+str(ArrearToPay)+'\","0",\"'+str(creditStart)+'\",\"'+str(creditStart)+'\","0")' )
        conn.commit()



conn.close()
